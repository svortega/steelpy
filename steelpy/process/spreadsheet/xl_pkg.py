# Copyright (c) 2022 steelpy
#
# Python stdlib imports
# from dataclasses import dataclass
from pathlib import Path
import re
import os
from typing import NamedTuple, Dict, List, Iterable, Union

# package imports
from steelpy.process.spreadsheet.dataframe import DataFrame
from steelpy.process.spreadsheet.pylightxl.pylightxl import(utility_address2index,
                                                            utility_columnletter2num,
                                                            utility_index2address)


#
class ExcelExt:
    __slots__ = ['units', '_wb', '_ws', '_flag']

    def __init__(self, wb_name: str):
        """
        """
        self._flag = True
        try:
            import xlwings as xw
            xw.Book(wb_name).set_mock_caller()
            self._wb = xw.Book.caller()
            self._flag = False
        except ModuleNotFoundError:
            try:
                from openpyxl import load_workbook
                #self._wb = load_workbook(wb_name, read_only=False,
                #                         keep_vba=True, data_only=True)
                self._wb = load_workbook(wb_name)
            except PermissionError:
                raise PermissionError('close woorkbook or install xlwings')
        #
        #
        self._ws = Sheets(self._wb)
        # print('-->')

    #
    @property
    def sheets(self):
        """ """
        return self._ws    
    #
    #
    #
    @property
    def sheet_names(self):
        """ """
        return self._ws.ws_names  
    #
    #
    def save(self):
        """ """
        if self._flag:
            self._ws.close()


#
#
class Sheets:
    __slots__ = ['units', '_wb', '_cells']

    def __init__(self, wb: str):
        """
        """
        self._wb = wb
    #
    def __getitem__(self, ws_name: str):
        """
        """
        s_name = self.ws_names
        if ws_name in s_name:
            return xlCells(self._wb, ws_name)
        else:
            raise IOError(f"sheet {ws_name} not found")
    #
    @property
    def ws_names(self):
        """ get sheet names"""
        try:
            return [sh.name for sh in self._wb.sheets]
        except AttributeError:
            return self._wb.sheetnames
    #
    def close(self):
        """ """
        #if self._flag:
        self._wb.close()    
#
#
#
class xlCells:
    __slots__ = ['_wb', '_ws_name', '_ws']

    def __init__(self, wb, ws_name: str):
        """ ws : sheet"""
        self._wb = wb
        self._ws_name = ws_name
        try: # xlwings
            self._ws = self._wb.sheets[self._ws_name]
        except AttributeError: # openpyxl
            self._ws  = self._wb[self._ws_name]        
    #
    #
    def __setitem__(self, cell_name: str, value) -> None:
        """
        """
        # if re.search(r"\:", cell_name):
        # tokens = re.split(r'[:]', cell_name)
        # row_i, col_i = utility_address2index(tokens[0])
        # row_j, col_j = utility_address2index(tokens[1])
        # for i in range(row_i, row_j+1):
        #    for j in range(col_i, col_j+1):
        #        self._wb.ws(ws=self._ws_name).update_index(row=i, col=j, val=value)
        if isinstance(value, list):
            row_i, col_j = utility_address2index(cell_name)
            columns = len(value)
            if isinstance(value[0], list):
                rows =  len(value[0])
                for j in range(columns):  # columns
                    for i in range(rows):  # rows
                        print('---', i)
                        self._wb.ws(ws=self._ws_name).update_index(row=row_i + i,
                                                                   col=col_j + j,
                                                                   val=value[j][i])
            else:
                for j in range(columns):  # columns
                    self._wb.ws(ws=self._ws_name).update_index(row=row_i,
                                                               col=col_j + j,
                                                               val=value[j])
        else:
            self._wb.ws(ws=self._ws_name).update_address(address=cell_name, val=value)

    #
    def __getitem__(self, cell_name: str):
        """
        """
        try:
            #ws = self._wb.sheets[self._ws_name]
            return self._ws.range(cell_name).value
        except AttributeError:
            #ws = self._wb[self._ws_name ]
            return [[item.value for item in rows] 
                    for rows in self._ws[cell_name]]
    #
    @property
    def column(self):
        """ """
        return xlColumn(self._wb, self._ws_name)
    #
    @property
    def row(self):
        """ """
        return xlRow(self._wb, self._ws_name)     
    #
    #
    def key(self, row:Union[int,str,None]=None, 
            column:Union[int,str,None]=None):
        """ """
        ws_name = self._ws_name
        
        #if row:
        #    if isinstance(column, str):
        #        
        #
        #if column:
        #    if isinstance(column, str):
        return self._wb.ws(ws=self._ws_name).ssd(keyrows=row, keycols=column)
    #
    #
    @property
    def dataframe(self):
        """ """
        return DataFrameOps(ws=self._ws)
    #
    def get_data(self):
        """ """
        try:
            data = self._ws.used_range.value
        except AttributeError:
            data = get_column_data(self._ws)
        return data    
#
#
#
class DataFrameOps:
    __slots__ = ['_ws',  '_column']

    def __init__(self, ws):
        """ ws : sheet"""
        self._ws = ws
        self._column = xlColumn(self._ws)
    #
    #
    def __setitem__(self, cell_name: str, df) -> None:
        """
        """
        if re.search(r"\:", cell_name):
            pass
        else:
            # [row, col]
            index = utility_address2index(cell_name)
            for x, (key, value) in enumerate(df.items()):
                row = index[0]
                col = index[1] + x
                address = utility_index2address(row, col)
                #self._ws.range(address).value = key
                try:
                    self._ws[address].value = key
                except AttributeError:
                    self._ws[address] = key
                address = utility_index2address(row+1, col)
                self._column[address] = value
    #
    def __call__(self, row:int=1, column:Union[int,str]=1,
                  title:Union[int,List]=0):
        """ """
        data = self.get_data()
        columns = list(zip(*data))
        #
        if isinstance(title, (list, tuple)):
            print('---')
            1/0
        else:
            headers = {col[title]: col[title+1:] for col in columns}
        #
        #print('--')
        return DataFrame(headers)
    #
    #
    def get_data(self):
        """ """
        try:
            data = self._ws.used_range.value
        except AttributeError:
            data = get_column_data(self._ws)
        return data     
#
#
class WriteExcelData:
    __slots__ = ['_wb', '_wb_name', '_flag',
                 '_col2name']

    def __init__(self, wb_name):
        """
        """
        self._flag = True
        try:
            import xlwings as xw
            xw.Book(wb_name).set_mock_caller()
            self._wb = xw.Book.caller()
            self._flag = False
        except ModuleNotFoundError:
            from openpyxl import load_workbook
            self._wb = load_workbook(wb_name)
        #
        path = Path().absolute()
        path = str(path) + f"\\{wb_name}"
        self._wb_name = path
        #
        self._col2name = ColumnName()
    #
    #
    def sheet_active(self, ws_name: str):
        """ """
        try:
            return self._wb[ws_name]
        except TypeError:
            # return self._wb.sheets[ws_name]
            return xlset(ws=self._wb.sheets[ws_name])

    #
    #
    def cell_name(self, column: int, row: int):
        """ return cell name"""
        col = self._col2name.get_column_name(column)
        name = f"{col}{row}".upper()
        return name

    #
    #
    # def cell(self,row:int, column:int, value:Union[str,int,float]):
    #    """ """
    #
    #
    def plot(self, ws_name: str, fig, anchor: str, name: str):
        """ """
        try:
            sheet = self._wb[ws_name]
            from openpyxl.drawing.image import Image
            fig.savefig('unit.png')
            img = Image('unit.png')
            # img.width = 640
            # img.height = 780
            img.anchor = anchor  # "I50"
            sheet.add_image(img)
        except TypeError:
            sheet = self._wb.sheets[ws_name]
            sheet.pictures.add(fig, name=name, update=True)
            #

    def update_column(self, ws_name: str, col: int,
                      values: Dict, step: int):
        """ """
        try:  # openpyxl
            cells = self._wb[ws_name]
        except TypeError:  # xlwings
            ws = self._wb.sheets[ws_name]
            # cells = ws.cells
            cells = ws.range((1, 1), (135, 16))
        #
        for row in cells.rows:
            name = row[col].value
            try:
                for x, item in enumerate(values[name]):
                    row[step + x].value = item
            except KeyError:
                continue
        # print('--')

    #
    def close(self):
        """ """
        if self._flag:
            self._wb.save(self._wb_name)
        #


#
class NewExcelFile:
    __slots__ = ['_wb', '_wb_name']

    def __init__(self, wb_name: str):
        """
        """
        self._wb = Database()
        self._wb_name = wb_name
    #
    #
    def sheet_active(self, ws_name: str):
        """ """
        self._wb.add_ws(ws=ws_name)
        return xlsetNew(wb=self._wb, ws_name=ws_name)

    #
    def close(self, file_name):
        """ """
        #fn = self._wb_name + ".xlsx"
        #fn = os.path.join(self._path, fn)
        writexl(db=self._wb, fn=file_name)


#
class xlsetNew:
    __slots__ = ['_wb', '_ws_name']

    def __init__(self, wb, ws_name: str):
        """ ws : sheet"""
        self._wb = wb
        self._ws_name = ws_name

    #
    #
    def __setitem__(self, cell_name: str, value) -> None:
        """
        """
        # if re.search(r"\:", cell_name):
        # tokens = re.split(r'[:]', cell_name)
        # row_i, col_i = utility_address2index(tokens[0])
        # row_j, col_j = utility_address2index(tokens[1])
        # for i in range(row_i, row_j+1):
        #    for j in range(col_i, col_j+1):
        #        self._wb.ws(ws=self._ws_name).update_index(row=i, col=j, val=value)
        if isinstance(value, list):
            row_i, col_j = utility_address2index(cell_name)
            columns = len(value)
            if isinstance(value[0], list):
                rows =  len(value[0])
                for j in range(columns):  # columns
                    for i in range(rows):  # rows
                        print('---', i)
                        self._wb.ws(ws=self._ws_name).update_index(row=row_i + i, 
                                                                   col=col_j + j,
                                                                   val=value[j][i])                        
            else:
                for j in range(columns):  # columns
                    self._wb.ws(ws=self._ws_name).update_index(row=row_i, 
                                                               col=col_j + j,
                                                               val=value[j])
        else:
            self._wb.ws(ws=self._ws_name).update_address(address=cell_name, val=value)

    #
    def __getitem__(self, cell_name: str):
        """
        """
        if re.search(r"\:", cell_name):
            tokens = re.split(r'[:]', cell_name)
            row_i, col_i = utility_address2index(tokens[0])
            row_j, col_j = utility_address2index(tokens[1])
            for i in range(row_i, row_j + 1):
                for j in range(col_i, col_j + 1):
                    self._wb.ws(ws=self._ws_name).update_index(row=i, col=j, val=value)
        else:
            return self._wb.ws(ws=self._ws_name).index(address=cell_name)
            # row_id, col_id = utility_address2index(cell_name)
            # return self._wb.ws(ws=self._ws_name).index(row=row_id, col=col_id)


#
def get_column_data(sheet):
    """ """
    cell_obj = []
    for row in sheet.rows:
        cell_obj.append([])
        for cell in row:
            cell_obj[-1].append(cell.value)
    return cell_obj
#
#
class xlset:
    __slots__ = ['_ws']

    def __init__(self, ws):
        """ ws : sheet"""
        self._ws = ws

    #
    #
    def __setitem__(self, cell_name: str, value: float) -> None:
        """
        """
        self._ws.range(cell_name).value = value

    #
    def __getitem__(self, cell_name: int):
        """
        """
        return self._ws.range(cell_name).value


#
#
#class ColumnName(dict):
#    """
#    """
#
#    def __init__(self):
#        import string
#        super(ColumnName, self).__init__()
#        self.alphabet = string.ascii_uppercase
#        self.alphabet_size = len(self.alphabet)
#
#    def __missing__(self, column_number):
#        ret = self[column_number] = self.get_column_name(column_number)
#        return ret
#
#    def get_column_name(self, column_number):
#        if column_number <= self.alphabet_size:
#            return self.alphabet[column_number - 1]
#        else:
#            return self.alphabet[int(((column_number - 1) / self.alphabet_size)) - 1] + self.alphabet[
#                ((column_number - 1) % self.alphabet_size)]
#
#
#
#
def get_row_column(data_list: List, header: str):
    """ """
    row = [x for x, item in enumerate(data_list) if header in item]
    col = [x for x, item in enumerate(data_list[row[0]]) if item == header]
    return [row[0], col[0]]
#
#
#
class xlColumn:
    __slots__ = ['_ws']

    def __init__(self, ws):
        """ ws : sheet"""
        self._ws = ws
    #
    def __setitem__(self, col_name: str, value: float) -> None:
        """
        """
        if isinstance(col_name, int):
            1/0
        elif isinstance(col_name, str):
            numbers = re.findall('[0-9]+', col_name)
            if numbers:
                row, col = utility_address2index(col_name)
                row += 1
            else:
                col = utility_columnletter2num(col_name)
                row = 2
            rows = row + len(value)
            address = utility_index2address(rows, col)
            cell_name = col_name + f":{address}"
        else:
            raise IOError(f"column name : {col_name} not valid")
        #
        # xlwings
        try:
            self._ws.range(cell_name).options(transpose=True).value = value
        except: # openpyxl
            for x, item in enumerate(value):
                row += x
                self._ws.cell(row=row,column=col).value = item         
    #
    #
    def __getitem__(self, col_name: Union[int, str]):
        """
        """
        if isinstance(col_name, int):
            return self._wb(ws=self._ws_name).col(col=col_name)
        elif isinstance(col_name, str):
            numbers = re.findall('[0-9]+', col_name)
            if numbers:
                row, col = utility_address2index(col_name)
                1/0
            else:
                #col_name = utility_columnletter2num(col_name)
                #row = 1
                return self._wb.ws(ws=self._ws_name).col(col=col_name)
        else:
            raise IOError(f"column name : {col_name} not valid")
#
#
#
